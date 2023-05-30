import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from dataclasses import dataclass
from openpyxl.styles.alignment import Alignment
from typing import List, Union, Optional, Dict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, Border, Side, PatternFill
from enum import Enum


class Type(Enum):
    TEXT = "text"
    CATEGORY = "category"
    FLOAT = "float"
    NUMBER = "number"
    BOOLEAN = "boolean"


@dataclass
class ExpectedColumnItem:
    name: str
    procedure: str
    type: Type
    required: bool = True
    options: Optional[List[str]] = None
    regex: Optional[re.Pattern] = None
    min: Optional[Union[int, float]] = None
    max: Optional[Union[int, float]] = None
    description: Optional[str] = None
    example: Optional[Union[str, int, float]] = None


class BaseSpec:
    def __init__(self) -> None:
        pass

    @property
    def version(self) -> str:
        raise NotImplementedError

    @property
    def description(self) -> str:
        raise NotImplementedError

    @property
    def specs(self) -> Dict[str, List[ExpectedColumnItem]]:
        raise NotImplementedError

    def _color_generator(self):
        """Yield color based on color map from https://colorbrewer2.org/#type=qualitative&scheme=Paired&n=9"""
        colors = [
            "A6CEE3",
            "1F78B4",
            "B2DF8A",
            "33A02C",
            "FB9A99",
            "E31A1C",
            "FDBF6F",
            "FF7F00",
            "CAB2D6",
            "6A3D9A",
            "FFFF99",
            "B15928",
        ]
        for color in colors:
            yield color

    def _gen_colors(self) -> Dict[str, List[str]]:
        """Generate colors for each procedure"""
        colors = {}
        for sheet in self.specs.keys():
            procedures = [
                item.procedure
                for item in self.specs[sheet]
                if item.procedure is not None
            ]

            color_map = {}
            color = self._color_generator()
            for procedure in procedures:
                if procedure not in color_map:
                    color_map[procedure] = next(color)

            colors[sheet] = [color_map[procedure] for procedure in procedures]

        return colors

    def _auto_width(self, ws):
        # 设置自动列宽
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[
                get_column_letter(column[0].column)
            ].width = adjusted_width

        return ws

    def _auto_height(self, ws):
        # 设置自动行高
        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                try:
                    if cell.value:
                        cell_height = (
                            len(str(cell.value).split("\n")) * 14
                        )  # 14是大概的行高值，可以根据需要调整
                        if cell_height > max_height:
                            max_height = cell_height
                except:
                    pass
            ws.row_dimensions[row[0].row].height = max_height

    def _max_width(self, ws, max_width=45):
        for column in ws.columns:
            ws.column_dimensions[get_column_letter(column[0].column)].width = max_width

    def _border(self, ws):
        # 设置线框样式
        border_style = Side(style="thin")
        border = Border(
            top=border_style, right=border_style, bottom=border_style, left=border_style
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    def _align(self, ws):
        alignment = Alignment(wrapText=True, vertical="center", horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

    def generate_template(self, filepath: Path) -> Workbook:
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)  # type: ignore

        ws = wb.create_sheet("Sheet1")

        ws.title = "Please Read First!"
        ws.merge_cells("A1:F1")  # type: ignore

        # Set the color as red
        font = Font(name="Arial", size=22, bold=True, color=Color(rgb="FF0000"))
        ws["A1"] = "Please Read First (%s)" % self.version
        ws["A1"].font = font

        ws.merge_cells("A2:F3")
        ws["A2"] = self.description

        for sheet in self.specs.keys():
            # Get current row number
            row = ws.max_row + 2
            ws.merge_cells("A%d:F%d" % (row, row))
            ws["A%d" % row] = "Sheet - %s" % sheet.capitalize()

            d = [
                {
                    "key": item.name,
                    "description": item.description,
                    "procedure": item.procedure,
                    "type": item.type.value,
                    "required": "Yes" if item.required else "No",
                    "example": item.example,
                }
                for item in self.specs[sheet]
            ]

            keys = ["key", "description", "procedure", "type", "required", "example"]
            # Write the column names (keys from the first dictionary) to the first row
            current_row = ws.max_row + 1
            for j, key in enumerate(keys, start=1):
                ws.cell(row=current_row, column=j, value=key)

            current_row = ws.max_row + 1
            colors = self._gen_colors()
            # Write the data from each dictionary to the following rows
            for j, item in enumerate(d, start=current_row):  # start from the second row
                color = colors[sheet][j - current_row]
                fill = PatternFill(fgColor=Color(rgb=color), fill_type="lightGray")
                for i, key in enumerate(item.keys(), start=1):
                    cell = ws.cell(row=j, column=i, value=item[key])
                    cell.fill = fill

            # Add a blank row
            ws.append([])

        # ws = self._auto_width(ws)
        # Set the max width for description column
        self._max_width(ws, max_width=30)
        self._align(ws)
        self._border(ws)

        # Add a new worksheet
        for sheet in self.specs.keys():
            ws = wb.create_sheet(sheet)
            columns = [item.name for item in self.specs[sheet]]
            ws.append(columns)

            examples = [item.example for item in self.specs[sheet]]
            ws.append(examples)

            self._max_width(ws, max_width=30)
            self._align(ws)

        # Write to file
        wb.save(filepath)

        return wb

    @property
    def sheet_names(self) -> List[str]:
        return list(self.specs.keys())
